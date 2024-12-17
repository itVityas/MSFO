import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from msfo1.models import Debt, AccountMapping
from django.conf import settings
from datetime import datetime


def set_column_widths(ws):
    """
    Устанавливает ширину столбцов, высоту строки с заголовками
    """
    column_widths = {
        'A': 97.14,
        'B': 12.14,
        'C': 18.43,
        'D': 16.00,
        'E': 13.71,
        'F': 16.71,
        'G': 18.29,
        'H': 17.29,
        'I': 14.71,
        'J': 80.57,
        'K': 15.29,
        'L': 45.14,
        'M': 14.71,
        'N': 25.86,
        'O': 16.29,
        'P': 16.57,
        'Q': 16.57
    }

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[2].height = 30


def set_font(ws):
    """
    Устанавливает шрифт на всем листе
    """
    font = Font(name='Arial', size=9)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            cell.font = font


def set_headers(ws, report_date):
    """
    Заполняет заголовки
    """
    headers = [
        "Контрагент",
        "Счет БСУ",
        "Задолженность в белорусских рублях",
        "Задолженность в валюте договора",
        "Валюта договора",
        "Дата возникновения задолженности",
        "Контрактные сроки погашения задолженности",
        "Контрактные сроки погашения задолженности",
        "Количество дней просрочки",
        "Примечание",
        "Счет МСФО",
        "Характер задолженности",
        "Курс валюты на 31.12.2023",
        "Задолженность в белорусских рублях по курсу 31.12.2023",
        "Курсовые разницы",
        "% резервирования",
        "Сумма резерва по номиналу"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    report_date = datetime(report_date, 12, 31)
    cell_data = ws.cell(row=1, column=9, value=report_date)
    cell_data.number_format = 'DD.MM.YYYY'


def highlight_cells(ws, end_row):
    """
    Закрашивает ячейки C, D, F, G, H с 3-ей строки по end_row,
    а также всю строку end_row с A по Q.
    """
    # Определяем заливку (желтый цвет: #FFFF00)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Столбцы: A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, ..., Q=17
    columns_to_fill = [3, 4, 6, 7, 8]  # C, D, F, G, H

    # Закрашиваем C, D, F, G, H с 3 строки по end_row
    for row in range(3, end_row + 1):
        for col in columns_to_fill:
            cell = ws.cell(row=row, column=col)
            cell.fill = yellow_fill

    # Закрашиваем строку end_row с A по Q
    for col in range(1, 18):  # 1 до 17 включительно
        cell = ws.cell(row=end_row, column=col)
        cell.fill = yellow_fill


def set_all_borders(ws, end_row):
    """
    Устанавливает тонкие границы для всех ячеек с 2 строки по current_row,
    от столбца A (1) до Q (17).
    """
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in range(2, end_row + 1):
        for col in range(1, 18):  # 1 - A, 17 - Q
            cell = ws.cell(row=row, column=col)
            cell.border = border


def set_numeric_format(ws, end_row):
    """
    Устанавливает числовой формат для столбцов C, D, N, O, Q
    с 3 по current_row.
    """
    numeric_columns = [3, 4, 14, 15, 17]  # C=3, D=4, N=14, O=15, Q=17
    for row in range(3, end_row + 1):
        for col in numeric_columns:
            cell = ws.cell(row=row, column=col)
            cell.number_format = '# ### ### ##0.00'


def fill_data_for_account_number(ws, db_account_number, report_file):
    """
    Заполнение .xlsx файла данными из бд, формулами
    """
    debts = Debt.objects.filter(
        account__db_account_number=db_account_number,
        report_file=report_file
    ).order_by('counterparty__name')

    current_row = 3

    # Если нет данных
    if not debts.exists():
        ws.cell(row=current_row, column=1, value="Нет данных по этому счету")
        current_row += 1
        return current_row

    # Заполняем таблицу данными
    for debt in debts:
        # Контрагент
        ws.cell(row=current_row, column=1, value=debt.counterparty.name)

        # Счет БСУ
        ws.cell(row=current_row, column=2, value=debt.account.db_account_number)

        # Задолженность в BYN
        ws.cell(row=current_row, column=3, value=debt.debt_byn)

        # Задолженность в валюте договора или в BYN, в зависимости от валюты
        if debt.contract_currency == 'BYN':
            ws.cell(row=current_row, column=4, value=debt.debt_byn)
        else:
            ws.cell(row=current_row, column=4, value=debt.debt_contract_currency)

        # Валюта договора
        ws.cell(row=current_row, column=5, value=debt.contract_currency)

        # Дата возникновения задолженности
        cell_data = ws.cell(row=current_row, column=6, value=debt.date_of_debt)
        cell_data.number_format = 'DD.MM.YYYY'

        # Контрактные сроки погашения задолженности
        ws.cell(row=current_row, column=7, value=debt.payment_term_days)

        # Формулы
        ws.cell(row=current_row, column=8, value=f"=F{current_row}+G{current_row}")
        ws.cell(row=current_row, column=9, value=f"=IF(H{current_row}>$I$1,0,$I$1-H{current_row})")
        ws.cell(row=current_row, column=10, value='')
        ws.cell(row=current_row, column=11, value=1.314)
        ws.cell(row=current_row, column=12, value='монетарная')
        ws.cell(row=current_row, column=13, value=f'=IF(E{current_row}="BYN",1,"см")')
        ws.cell(row=current_row, column=14, value=f"=M{current_row}*D{current_row}")
        ws.cell(row=current_row, column=15, value=f"=N{current_row}-C{current_row}")
        ws.cell(row=current_row, column=16, value=(
            f"=IF(OR(K{current_row}=1.314,K{current_row}=4.104),0,"
            f"IF(I{current_row}>365,1,VLOOKUP(I{current_row},'% резервирования'!$A$3:$B$369,2,0)))"
        ))
        ws.cell(row=current_row, column=17, value=f"=P{current_row}*N{current_row}")

        current_row += 1

    # Добавляем итоговые строки с суммами значений
    ws.cell(row=current_row, column=1, value='Итого:')
    ws.cell(row=current_row, column=3, value=f'=SUM(C3:C{current_row-1})')
    ws.cell(row=current_row, column=14, value=f'=SUM(N3:N{current_row-1})')
    ws.cell(row=current_row, column=15, value=f'=SUM(O3:O{current_row-1})')
    ws.cell(row=current_row, column=17, value=f'=SUM(Q3:Q{current_row-1})')

    # Автофильтр в ячейки заголовка
    ws.auto_filter.ref = f"A2:Q{current_row-1}"

    return current_row


def create_and_fill_ws(wb, year, db_account_number, report_file):
    """
    Создает лист и заполняет его данными
    """
    ws_name = f"{db_account_number}-{year}"
    ws = wb.create_sheet(title=ws_name)

    set_headers(ws=ws, report_date=year)
    end_row = fill_data_for_account_number(ws=ws, db_account_number=db_account_number, report_file=report_file)
    set_font(ws=ws)
    set_column_widths(ws=ws)
    highlight_cells(ws=ws, end_row=end_row)
    set_all_borders(ws=ws, end_row=end_row)
    set_numeric_format(ws=ws, end_row=end_row)
    print(f'ws {ws_name} was successfully created.')


def generate_msfo_report(year, report_file):
    """
    Создает полный отчет мсфо
    """
    source_file = os.path.join(settings.BASE_DIR, 'msfo1', 'static', 'source', 'msfo1.xlsx')
    wb = load_workbook(source_file)

    # Получаем уникальные db_account_number по возрастанию
    db_accounts = (
        AccountMapping.objects.values_list('db_account_number', flat=True)
        .distinct()
        .order_by('db_account_number')
    )

    for db_account_number in db_accounts:
        create_and_fill_ws(wb, year, db_account_number, report_file)


    # Сохраняем файл
    file_name = datetime.now().strftime(f"Adjustment 1 - {year}. Created at %Y.%m.%d %H:%M:%S.xlsx")
    output_file = os.path.join(settings.BASE_DIR, 'msfo1', 'static', 'xlsx', file_name)
    wb.save(output_file)

    # Сохраняем запись в БД
    report_file.file_path = output_file
    report_file.save()

    print(f"Отчет сформирован и сохранен по пути: {output_file}")

    # Удаляем все данные, относящиеся к определенному report_file
    Debt.objects.filter(report_file=report_file).delete()
