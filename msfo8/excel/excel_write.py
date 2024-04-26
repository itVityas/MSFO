import datetime
from openpyxl import load_workbook, Workbook, utils
from openpyxl.styles import PatternFill, Alignment, Font, numbers
from msfo8.models import Entrance, EGIL
from msfo8.excel.utils import get_report


def create_ig2014():
    workbook = Workbook()
    sheet = workbook.active
    headers = ['Data', 'Month index', 'Year index', 'Start hyperinflation index', 'Index for hyperinflation']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    egil_objects = EGIL.objects.all()
    for row_num, egil_object in enumerate(egil_objects, 2):
        sheet.cell(row=row_num, column=1, value=egil_object.data)
        sheet.cell(row=row_num, column=2, value=egil_object.month_index)
        sheet.cell(row=row_num, column=3, value=egil_object.year_index)
        sheet.cell(row=row_num, column=4, value=egil_object.start_hyper_index)
        sheet.cell(row=row_num, column=5, value=egil_object.hyper_index)
    workbook.save('/home/foile/MSFO/MSFO/static/xlsx/egil_data.xlsx')


def entrance_create(line):
    price = f'=J{line}/F{line}'
    write_off = f'=IF(G{line}<$L$1,"списываем","не списываем")'
    reclass = f'=IF(L{line}="списываем",4.104,1.403)'
    necessity_reserve = f'=IF(AND($M$1>G{line},L{line}="не списываем"),"да","нет")'
    ig2014 = f'=IF($N$1<G{line},1,VLOOKUP(G{line},ИГ2014!$G$7:$H$295,2,1))'
    cost_msfo = f'=IF(L{line}="списываем",0,N{line}*J{line})'
    write_up = f'=IF(L{line}="не списываем",O{line}-J{line},0)'
    cost_write_off = f'=IF(L{line}="списываем",J{line}-O{line},0)'
    reserve = f'=IF(M{line}="да",O{line},0)'
    return price, write_off, reclass, necessity_reserve, ig2014, cost_msfo, write_up, cost_write_off, reserve


def change_column_wight(wb_list):
    column_widths = {
        'A': 21.15,
        'B': 13.43,
        'C': 12.43,
        'D': 17.85,
        'E': 12.71,
        'F': 10.43,
        'G': 16.28,
        'H': 62.71,
        'I': 8.53,
        'J': 15.57,
        'K': 8.53,
        'L': 12.85,
        'M': 15.0,
        'N': 13.71,
        'O': 14.85,
        'P': 17.71,
        'Q': 14.57,
        'R': 13.43
    }

    for column, width in column_widths.items():
        wb_list.column_dimensions[column].width = width
    wb_list.row_dimensions[2].height = 30


def change_font(wb_list):
    font = Font(name='Arial', size=9)
    for row in wb_list.iter_rows():
        for cell in row:
            cell.font = font


def change_number_format(wb_list, end_row):
    columns = [5, 6, 10, 15, 16, 17, 18]
    for row in range(3, end_row + 1):
        for column in columns:
            cell = wb_list.cell(row=row, column=column)
            cell.number_format = numbers.FORMAT_NUMBER_00
    return


def write_header(wb_list, report_name):

    headers = ['Дата остатков', 'Склад', 'Счет', 'Номенклатурный №', 'Цена', 'Кол-во',
               'Дата поступления', 'Наименование', 'Ед.изм.', 'Стоимость', 'Счет рекласса МСФО',
               'Списание запасов', 'Необходимость формирования резерва', 'ИГ 2014',
               'Стоимость МСФО на 31.12.2021', 'Нереализованная ГИ дооценка',
               'Стоимость списанных материалов', 'Резерв МСФО']

    for col_num, header in enumerate(headers, 1):
        cell = wb_list.cell(row=2, column=col_num, value=header, )
        # cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

    id_report, date_write_off, date_necessity, date_ig2014 = get_report(report_name)

    wb_list.cell(row=1, column=12, value=date_write_off)
    wb_list.cell(row=1, column=13, value=date_necessity)
    wb_list.cell(row=1, column=14, value=date_ig2014)


def write_data(wb_list, date, num_store):
    entrances = Entrance.objects.select_related('id_material', 'id_report',
                                                'id_bill', 'id_store').filter(id_bill__number=num_store)

    line = 3
    for row_num, entrance in enumerate(entrances, 3):
        (price, write_off, reclass, necessity_reserve, ig2014, cost_msfo, write_up, cost_write_off,
         reserve) = entrance_create(line)
        line += 1
        wb_list.cell(row=row_num, column=1, value=date)
        wb_list.cell(row=row_num, column=2, value=entrance.id_store.numbers)
        wb_list.cell(row=row_num, column=3, value=entrance.id_bill.number)
        wb_list.cell(row=row_num, column=4, value=entrance.id_material.code)
        wb_list.cell(row=row_num, column=5, value=price)
        wb_list.cell(row=row_num, column=6, value=entrance.count)
        wb_list.cell(row=row_num, column=7, value=entrance.date)
        wb_list.cell(row=row_num, column=8, value=entrance.id_material.name)
        wb_list.cell(row=row_num, column=9, value=entrance.id_material.measuring)
        wb_list.cell(row=row_num, column=10, value=entrance.all_price)
        wb_list.cell(row=row_num, column=11, value=reclass)
        wb_list.cell(row=row_num, column=12, value=write_off)
        wb_list.cell(row=row_num, column=13, value=necessity_reserve)
        wb_list.cell(row=row_num, column=14, value=ig2014)
        wb_list.cell(row=row_num, column=15, value=cost_msfo)
        wb_list.cell(row=row_num, column=16, value=write_up)
        wb_list.cell(row=row_num, column=17, value=cost_write_off)
        wb_list.cell(row=row_num, column=18, value=reserve)
    wb_list.cell(row=line, column=8, value='Итого:')
    wb_list.cell(row=line, column=10, value=f'=SUM(J3:J{line-1})')
    wb_list.cell(row=line, column=15, value=f'=SUM(O3:O{line-1})')
    wb_list.cell(row=line, column=16, value=f'=SUM(P3:P{line-1})')
    wb_list.cell(row=line, column=17, value=f'=SUM(Q3:Q{line-1})')
    wb_list.cell(row=line, column=18, value=f'=SUM(R3:R{line-1})')
    change_number_format(wb_list, line)


def write_bill(workbook, num_bill, date, report_name):
    wb_list = workbook.create_sheet(f'{num_bill}-{date}')
    change_column_wight(wb_list)
    write_header(wb_list, report_name)
    write_data(wb_list, date, num_bill)
    change_font(wb_list)
    return


def write_all_date(date, report_name):
    workbook = load_workbook('/home/foile/MSFO/MSFO/static/IG2014.xlsx')
    write_bill(workbook, 1001, date, report_name)
    write_bill(workbook, 1002, date, report_name)
    datetime_now = datetime.datetime.now()
    wb_path = datetime_now.strftime(f'/home/foile/MSFO/MSFO/static/xlsx/%Y-%m-%d/data - %H:%M:%S.xlsx')
    workbook.save(wb_path)
    return wb_path


# write_all_date('2023', '2023')
