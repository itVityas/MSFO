from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment
import datetime
from msfo8.models import Bill, Store, Material, Entrance, Report, EGIL


class CellException(Exception):
    """Cell value is no correct"""
    pass


def get_or_create_bill(wb_list):
    name = wb_list['A14'].value
    number = name.replace('.', '0')
    id_bill, created = Bill.objects.get_or_create(name=name, number=number)
    return id_bill


def get_or_create_store(line: int, wb_list):
    name = wb_list[f'A{line}'].value
    numbers = wb_list[f'C{line}'].value
    if numbers is None:
        raise CellException
    id_store, created = Store.objects.get_or_create(name=name, numbers=numbers)
    return id_store


def create_report(name, date_write_off, date_necessity, date_ig2014):
    id_report, created = Report.objects.get_or_create(name=name, date_write_off=date_write_off,
                                                      date_necessity=date_necessity, date_ig2014=date_ig2014)
    return id_report


def get_report(name):
    report = Report.objects.get(name=name)
    id_report = report
    date_write_off = report.date_write_off
    date_necessity = report.date_necessity
    date_ig2014 = report.date_ig2014
    return id_report, date_write_off, date_necessity, date_ig2014


def read_material(line: int, wb_list):
    name = wb_list[f'A{line}'].value
    code = wb_list[f'C{line}'].value
    measuring = wb_list[f'D{line}'].value
    return name, code, measuring


def read_date(line: int, wb_list):
    date = wb_list[f'A{line}'].value
    date = date_convert(date)                       # Index error if not correct cell
    all_price = wb_list[f'H{line}'].value
    if all_price is None:
        all_price = 0
    count = wb_list[f'H{line + 1}'].value
    if count is None:
        raise CellException
    return date, all_price, count


def date_convert(date: str):
    date_list = date.split('.')
    date = datetime.date(int(date_list[2]), int(date_list[1]), int(date_list[0]))
    return date


def entrance_create(line):
    price = f'=J{line}/F{line}'
    write_off = f'=IF(G{line}<$L$1,"списываем","не списываем")'
    reclass = f'=IF(L{line}="списываем",4.104,1.403)'
    necessity_reserve = f'=IF(AND($M$1>G{line},L{line}="не списываем"),"да","нет")'
    ig2014 = f'=IF($N$1<G{line},1,VLOOKUP(G{line},$ИГ2014.$G$7:$H$295,2,1))'
    cost_msfo = f'=IF(L{line}="списываем",0,N{line}*J{line})'
    write_up = f'=IF(L{line}="не списываем",O{line}-J{line},0)'
    cost_write_off = f'=IF(L{line}="списываем",J{line}-O{line},0)'
    reserve = f'=IF(M{line}="да",O{line},0)'
    return price, write_off, reclass, necessity_reserve, ig2014, cost_msfo, write_up, cost_write_off, reserve


def write_all_date_bd(path, report_name):
    wb = load_workbook(f'{path}', data_only=True)
    wb_list = wb.active
    line = 16
    id_report, date_write_off, date_necessity, date_ig2014 = get_report(report_name)
    id_bill = get_or_create_bill(wb_list)
    while True:
        try:
            id_store = get_or_create_store(line, wb_list)
        except CellException:
            break
        line += 2
        while True:
            name, code, measuring = read_material(line, wb_list)
            if measuring is None:
                break
            id_material, created = Material.objects.get_or_create(name=name, code=code,
                                                                  measuring=measuring)
            line += 2
            flag_material = True
            while flag_material:
                try:
                    date, all_price, count = read_date(line, wb_list)
                    line += 2
                    Entrance.objects.create(id_material=id_material, id_report=id_report, id_bill=id_bill,
                                            id_store=id_store, date=date, all_price=all_price, count=count)
                except IndexError:
                    flag_material = False
                except ValueError:
                    flag_material = False
                except CellException:
                    line += 2
                    continue
    wb.close()
    return


def create_ig2014():
    workbook = Workbook()
    sheet = workbook.active
    headers = ['Data', 'Month index', 'Year index', 'Start hyperinflation index', 'Index for hyperinflation']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    egil_objects = EGIL.objects.all()
    for row_num, egil_object in enumerate(egil_objects, 2):
        sheet.cell(row=row_num, column=1, value=egil_object.data)
        sheet.cell(row=row_num, column=2, value=egil_object.month_index)
        sheet.cell(row=row_num, column=3, value=egil_object.year_index)
        sheet.cell(row=row_num, column=4, value=egil_object.start_hyper_index)
        sheet.cell(row=row_num, column=5, value=egil_object.hyper_index)
    workbook.save('/home/foile/MSFO/MSFO/static/xlsx/egil_data.xlsx')


def write_all_date_xlsx(name_list, date, report_name):
    workbook = load_workbook('/home/foile/MSFO/MSFO/static/xlsx/IG2014.xlsx')
    wb_list = workbook.create_sheet(f'{name_list}')
    change_column_wight(wb_list)
    write_header(wb_list, report_name)
    write_data(wb_list, date)
    workbook.save('/home/foile/MSFO/MSFO/static/xlsx/data.xlsx')
    return


def change_column_wight(wb_list):
    column_widths = {
        'A': 21.15,
        'B': 13.43,
        'C': 12.43,
        'D': 17.85,
        'E': 12.71,
        'F': 10.43,
        'G': 16.28,
        'H': 62.71,
        'I': 8.53,
        'J': 15.57,
        'K': 8.53,
        'L': 12.85,
        'M': 15.0,
        'N': 13.71,
        'O': 14.85,
        'P': 17.71,
        'Q': 14.57,
        'R': 13.43
    }

    for column, width in column_widths.items():
        wb_list.column_dimensions[column].width = width


def write_header(wb_list, report_name):

    headers = ['Дата остатков', 'Склад', 'Счет', 'Номенклатурный №', 'Цена', 'Кол-во',
               'Дата поступления', 'Наименование', 'Ед.изм.', 'Стоимость', 'Счет рекласса МСФО',
               'Списание запасов', 'Необходимость формирования резерва', 'ИГ 2014',
               'Стоимость МСФО на 31.12.2021', 'Нереализованная ГИ дооценка',
               'Стоимость списанных материалов', 'Резерв МСФО']

    for col_num, header in enumerate(headers, 1):
        wb_list.cell(row=2, column=col_num, value=header)
        wb_list.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        wb_list.alignment = Alignment(horizontal="center")

    id_report, date_write_off, date_necessity, date_ig2014 = get_report(report_name)

    wb_list.cell(row=1, column=12, value=date_write_off)
    wb_list.cell(row=1, column=13, value=date_necessity)
    wb_list.cell(row=1, column=14, value=date_ig2014)


def write_data(wb_list, date):
    entrances = Entrance.objects.select_related('id_material', 'id_report', 'id_bill', 'id_store')

    line = 3
    for row_num, entrance in enumerate(entrances, 3):
        (price, write_off, reclass, necessity_reserve, ig2014, cost_msfo, write_up, cost_write_off,
         reserve) = entrance_create(line)
        line += 1
        wb_list.cell(row=row_num, column=1, value=date)
        wb_list.cell(row=row_num, column=2, value=entrance.id_store.numbers)
        wb_list.cell(row=row_num, column=3, value=entrance.id_bill.number)
        wb_list.cell(row=row_num, column=4, value=entrance.id_material.code)
        wb_list.cell(row=row_num, column=5, value=price)
        wb_list.cell(row=row_num, column=6, value=entrance.count)
        wb_list.cell(row=row_num, column=7, value=entrance.date)
        wb_list.cell(row=row_num, column=8, value=entrance.id_material.name)
        wb_list.cell(row=row_num, column=9, value=entrance.id_material.measuring)
        wb_list.cell(row=row_num, column=10, value=entrance.all_price)
        wb_list.cell(row=row_num, column=11, value=reclass)
        wb_list.cell(row=row_num, column=12, value=write_off)
        wb_list.cell(row=row_num, column=13, value=necessity_reserve)
        wb_list.cell(row=row_num, column=14, value=ig2014)
        wb_list.cell(row=row_num, column=15, value=cost_msfo)
        wb_list.cell(row=row_num, column=16, value=write_up)
        wb_list.cell(row=row_num, column=17, value=cost_write_off)
        wb_list.cell(row=row_num, column=18, value=reserve)


create_report(name='2023', date_write_off=datetime.date(2013, 12, 31),
              date_necessity=datetime.date(2020, 12, 31),
              date_ig2014=datetime.date(2015, 1, 1))

# Material.objects.all().delete()
# Store.objects.all().delete()
# write_all_date_bd('/home/foile/MSFO/MSFO/static/xlsx/test1001.xlsx', '2023')
write_all_date_xlsx('10-2023', '2023', '2023')
